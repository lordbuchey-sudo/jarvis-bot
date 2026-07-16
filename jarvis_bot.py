#!/usr/bin/env python3

"""
J.A.R.V.I.S. TELEGRAM BOT - FULL FEATURED
Unlimited Memory | Image Analysis | Voice Transcription
Premium Plans | Calendar | Email | CSV Generation
"""

import os
import io
import json
import re
import requests
import tempfile
import csv
import base64
from datetime import datetime, timedelta
from collections import defaultdict
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from telegram.request import HTTPXRequest

# ============================================
# API KEYS (reads from environment on Render, uses defaults locally)
# ============================================
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN", "YOUR_TELEGRAM_TOKEN")
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "YOUR_GROQ_KEY")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY", "YOUR_OPENROUTER_KEY")
ADMIN_USER_ID = int(os.getenv("ADMIN_USER_ID", "123456789"))

# ============================================
# PAYMENT LINKS (Replace with your own)
# ============================================
PAYMENT_LINKS = {
    "basic": "https://buy.stripe.com/your_basic_link",
    "business": "https://buy.stripe.com/your_business_link",
    "enterprise": "https://buy.stripe.com/your_enterprise_link"
}
PREMIUM_PRICES = {"basic": 49, "business": 149, "enterprise": 299}
premium_users = set()

# ============================================
# MEMORY SYSTEM
# ============================================
conversation_memory = defaultdict(list)
user_preferences = defaultdict(dict)
MAX_CONTEXT_MESSAGES = 20

def add_to_memory(user_id, role, content):
    conversation_memory[user_id].append({"role": role, "content": content})
    if len(conversation_memory[user_id]) % 5 == 0:
        save_memory_to_disk()

def get_recent_memory(user_id, count=MAX_CONTEXT_MESSAGES):
    if user_id not in conversation_memory or not conversation_memory[user_id]:
        return ""
    recent = conversation_memory[user_id][-count:]
    history = ""
    for msg in recent:
        role = "User" if msg["role"] == "user" else "Jarvis"
        history += f"{role}: {msg['content']}\n"
    return history

def get_all_memory(user_id):
    if user_id not in conversation_memory or not conversation_memory[user_id]:
        return "No messages yet."
    history = ""
    for msg in conversation_memory[user_id]:
        role = "You" if msg["role"] == "user" else "Jarvis"
        preview = msg['content'][:80] + "..." if len(msg['content']) > 80 else msg['content']
        history += f"{role}: {preview}\n"
    return history

def save_memory_to_disk():
    try:
        os.makedirs("data", exist_ok=True)
        serializable_memory = {}
        for user_id, messages in conversation_memory.items():
            serializable_memory[str(user_id)] = messages[-500:]
        with open("data/conversation_memory.json", "w") as f:
            json.dump(serializable_memory, f)
        serializable_prefs = {}
        for user_id, prefs in user_preferences.items():
            serializable_prefs[str(user_id)] = dict(prefs)
        with open("data/user_preferences.json", "w") as f:
            json.dump(serializable_prefs, f)
        with open("data/premium_users.json", "w") as f:
            json.dump(list(premium_users), f)
    except Exception as e:
        print(f"⚠️ Could not save memory: {e}")

def load_memory_from_disk():
    global premium_users
    try:
        if os.path.exists("data/conversation_memory.json"):
            with open("data/conversation_memory.json", "r") as f:
                data = json.load(f)
                for user_id, messages in data.items():
                    conversation_memory[int(user_id)] = messages
            print(f"✅ Loaded conversation memory from disk")
        if os.path.exists("data/user_preferences.json"):
            with open("data/user_preferences.json", "r") as f:
                data = json.load(f)
                for user_id, prefs in data.items():
                    user_preferences[int(user_id)] = prefs
            print(f"✅ Loaded user preferences from disk")
        if os.path.exists("data/premium_users.json"):
            with open("data/premium_users.json", "r") as f:
                premium_users = set(json.load(f))
            print(f"✅ Loaded premium users from disk")
    except Exception as e:
        print(f"⚠️ Could not load memory: {e}")

def get_title(user_id):
    return user_preferences.get(user_id, {}).get("title", None)

def get_name(user_id):
    return user_preferences.get(user_id, {}).get("name", "")

def is_premium(user_id):
    return user_id in premium_users

# ============================================
# AI ENGINE
# ============================================

def build_system_prompt(user_id):
    title = get_title(user_id)
    name = get_name(user_id)
    recent_memory = get_recent_memory(user_id, MAX_CONTEXT_MESSAGES)
    total_messages = len(conversation_memory.get(user_id, []))
    
    if name and title == "Ma'am":
        user_info = f"The user is a LADY. Her NAME is '{name}'. ALWAYS call her 'Ma'am {name}' or 'Ma'am'. 'Ma'am' is a TITLE, not her name."
        title_rule = f"CRITICAL: This user is a LADY named {name}. ALWAYS address her as 'Ma'am {name}' or 'Ma'am'. The word 'Sir' is ABSOLUTELY FORBIDDEN."
    elif name and title == "Sir":
        user_info = f"The user's NAME is '{name}'. His TITLE is 'Sir'. Call him 'Sir {name}' or 'Sir'."
        title_rule = f"Address the user as 'Sir {name}' or 'Sir'."
    elif name:
        user_info = f"The user's NAME is '{name}'. Title is '{title}'."
        title_rule = f"Address them as '{title} {name}'."
    elif title == "Ma'am":
        user_info = "The user is a LADY. Her title is 'Ma'am'. You don't know her name yet."
        title_rule = "CRITICAL: This user is a LADY. Call her 'Ma'am'. The word 'Sir' is FORBIDDEN."
    elif title == "Sir":
        user_info = "The user is a GENTLEMAN. His title is 'Sir'."
        title_rule = "Address him as 'Sir'."
    else:
        user_info = "You don't know the user's gender or title yet. Stay neutral."
        title_rule = "Be neutral and respectful. Don't use 'Sir' or 'Ma'am' unless the user tells you their preference."
    
    return f"""You are Jarvis, an elite AI assistant. Professional, witty, and highly capable.

{user_info}
{title_rule}

IMPORTANT: 'Ma'am' is a TITLE (like Mr., Mrs., Dr.), NOT a name.
If you know the user's name, use it: e.g., 'Ma'am Grace'

You have memory of our entire conversation ({total_messages} messages exchanged).
Refer to previous messages when relevant. Be consistent.

Current time: {datetime.now().strftime('%I:%M %p, %A %B %d, %Y')}

Recent conversation:
{recent_memory if recent_memory else 'New conversation'}"""

def ask_groq(user_id, message):
    url = "https://api.groq.com/openai/v1/chat/completions"
    headers = {"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"}
    messages = [{"role": "system", "content": build_system_prompt(user_id)}]
    for msg in conversation_memory.get(user_id, [])[-MAX_CONTEXT_MESSAGES:]:
        messages.append(msg)
    messages.append({"role": "user", "content": message})
    data = {"model": "llama-3.3-70b-versatile", "messages": messages, "temperature": 0.7, "max_tokens": 1000}
    response = requests.post(url, headers=headers, json=data, timeout=20)
    if response.status_code == 200:
        return response.json()["choices"][0]["message"]["content"]
    else:
        raise Exception(f"Groq error: {response.status_code}")

def ask_gemini_text(message):
    """Use Gemini for text responses"""
    if not GEMINI_API_KEY or GEMINI_API_KEY == "YOUR_GEMINI_KEY_HERE":
        raise Exception("Gemini not configured")
    
    try:
        from google import genai
        client = genai.Client(api_key=GEMINI_API_KEY)
        response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=message
        )
        return response.text
    except ImportError:
        raise Exception("Install: pip install google-genai")

def ask_openrouter(user_id, message):
    url = "https://openrouter.ai/api/v1/chat/completions"
    headers = {"Authorization": f"Bearer {OPENROUTER_API_KEY}", "Content-Type": "application/json", "HTTP-Referer": "https://jarvis-bot.local", "X-Title": "Jarvis Bot"}
    messages = [{"role": "system", "content": build_system_prompt(user_id)}]
    for msg in conversation_memory.get(user_id, [])[-MAX_CONTEXT_MESSAGES:]:
        messages.append(msg)
    messages.append({"role": "user", "content": message})
    data = {"model": "google/gemini-2.0-flash-lite-001:free", "messages": messages, "temperature": 0.7, "max_tokens": 1000}
    response = requests.post(url, headers=headers, json=data, timeout=25)
    if response.status_code == 200:
        return response.json()["choices"][0]["message"]["content"]
    else:
        raise Exception(f"OpenRouter error: {response.status_code}")

def jarvis_think(user_id, message):
    try:
        result = ask_groq(user_id, message)
        print("✅ Groq")
        return result
    except Exception as e:
        print(f"❌ Groq: {e}")
    try:
        result = ask_gemini_text(message)
        print("✅ Gemini")
        return result
    except Exception as e:
        print(f"❌ Gemini: {e}")
    try:
        result = ask_openrouter(user_id, message)
        print("✅ OpenRouter")
        return result
    except Exception as e:
        print(f"❌ OpenRouter: {e}")
    return f"⚠️ AI systems temporarily unavailable, {get_title(user_id)}. Please try again."

# ============================================
# USER INFO DETECTION
# ============================================

def detect_user_info(message, user_id):
    msg_lower = message.lower().strip()
    detected = False
    
    lady_triggers = [
        "not sir", "i'm madam", "i am madam", "call me madam",
        "stop calling me sir", "don't call me sir", "do not call me sir",
        "i'm female", "i am female", "i'm a woman", "i am a woman",
        "forget about the sir", "quit saying sir", "stop saying sir",
        "i'm a lady", "i am a lady", "im a lady",
        "why address me as sir", "why do you call me sir",
        "i'm not sir", "i am not sir", "im not sir",
        "sir is for gentlemen", "sir is for men", "i am not a man",
        "you cant call me sir", "can't call me sir", "cannot call me sir",
        "naa sir", "im lady", "i'm lady", "i am lady"
    ]
    
    for phrase in lady_triggers:
        if phrase in msg_lower:
            user_preferences[user_id]["title"] = "Ma'am"
            print(f"✅ Title set to MA'AM for user {user_id}")
            detected = True
            break
    
    if not detected:
        sir_triggers = ["not madam", "not ma'am", "i'm sir", "i am sir", "call me sir", "i'm male", "i am male", "i'm a man", "i am a man"]
        for phrase in sir_triggers:
            if phrase in msg_lower:
                user_preferences[user_id]["title"] = "Sir"
                print(f"✅ Title set to SIR for user {user_id}")
                detected = True
                break
    
    not_names = ["a","the","not","your","female","male","lady","woman","man","girl","boy","guy","madam","sir","ma'am","here","there","fine","ok","okay","good","bad","sure","ready","done","back","me","you","that","this","what","dont","don't","can't","cannot","just","add","call","saying","get","wat","meaning","arh","iu","naa","gentlemen","for","is","it","no","yes","why","how","who","where","when","which","am","are","was","were"]
    
    match = re.search(r"my name is\s+([a-zA-Z]+)", msg_lower)
    if match and match.group(1).lower() not in not_names:
        name = match.group(1)[0].upper() + match.group(1)[1:]
        user_preferences[user_id]["name"] = name
        print(f"✅ Name set to '{name}' for user {user_id}")
        detected = True
    
    match = re.search(r"(?:i'm|im|i am)\s+([a-zA-Z]+)\s+(?:who|and|that)\s+(?:is|i'm|im|i am)\s+(?:a\s+)?(?:lady|woman|girl|female)", msg_lower)
    if match and match.group(1).lower() not in not_names:
        name = match.group(1)[0].upper() + match.group(1)[1:]
        user_preferences[user_id]["name"] = name
        user_preferences[user_id]["title"] = "Ma'am"
        print(f"✅ Name set to '{name}' (lady pattern) for user {user_id}")
        detected = True
    
    if len(msg_lower.split()) <= 5:
        match = re.match(r"(?:i'm|im|i am)\s+([a-zA-Z]+)$", msg_lower)
        if match and match.group(1).lower() not in not_names:
            name = match.group(1)[0].upper() + match.group(1)[1:]
            user_preferences[user_id]["name"] = name
            print(f"✅ Name set to '{name}' (short intro) for user {user_id}")
            detected = True
    
    if detected:
        save_memory_to_disk()
    return detected

def is_affirmation(message):
    msg = message.lower().strip()
    affirmations = ["yes", "yeah", "yep", "yup", "sure", "ok", "okay", "go on", "continue", "tell me more", "elaborate", "dive deeper", "more", "yes please", "please do"]
    return msg in affirmations or msg.rstrip(".,!?") in affirmations

# ============================================
# USER DATA
# ============================================
user_data = {}

def get_user_data(user_id):
    if user_id not in user_data:
        user_data[user_id] = {"reminders": [], "notes": [], "documents_analyzed": 0, "joined": datetime.now().strftime("%Y-%m-%d %H:%M")}
    return user_data[user_id]

# ============================================
# COMMANDS
# ============================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    user_id = user.id
    title = get_title(user_id)
    name = get_name(user_id)
    greeting = "morning" if datetime.now().hour < 12 else "afternoon" if datetime.now().hour < 17 else "evening"
    welcome_line = f"Good {greeting}, {title} {name}." if name else f"Good {greeting}, {user.first_name}."
    total = len(conversation_memory.get(user_id, []))
    premium_badge = "💎 PREMIUM | " if is_premium(user_id) else ""
    await update.message.reply_text(
        f"🤖 *J.A.R.V.I.S. ONLINE*\n\n{welcome_line}\n\n{premium_badge}🧠 {total} messages stored\n📄 Docs | 📸 Images | 🎤 Voice | 📊 CSV | 📧 Email | 🗓 Calendar\n\n/help for all commands",
        parse_mode='Markdown'
    )

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "*📚 Commands:*\n\n"
        "/start — Welcome\n/help — This menu\n/status — AI status\n/whoami — What I know\n/history — Chat history\n"
        "/remind [time] [task] — Set reminder\n/note [text] — Save note\n"
        "/calc [math] — Calculate\n/search [query] — Search web\n"
        "/calendar add [event] — Add to calendar\n/calendar list — View events\n"
        "/email [to] [subject] [body] — Draft email\n"
        "/premium — View plans\n/buy [plan] — Get payment link\n/features — Your features\n"
        "/clear — Clear data\n/forget — Clear memory\n\n"
        "*Send:* 📸 Photo → Image analysis | 🎤 Voice → Transcription | 📄 File → Document analysis",
        parse_mode='Markdown'
    )

async def whoami_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    prefs = user_preferences.get(user_id, {})
    total = len(conversation_memory.get(user_id, []))
    premium = "✅ Yes" if is_premium(user_id) else "❌ No"
    await update.message.reply_text(
        f"👤 Title: {prefs.get('title', 'Not set')}\n📛 Name: {prefs.get('name', 'Not set')}\n🧠 Messages: {total}\n💎 Premium: {premium}\n💾 Saved: Yes",
        parse_mode='Markdown'
    )

async def history_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    history = get_all_memory(user_id)
    total = len(conversation_memory.get(user_id, []))
    if len(history) > 4000: history = history[:4000] + "\n...(truncated)"
    await update.message.reply_text(f"*📜 History ({total}):*\n\n{history}", parse_mode='Markdown')

async def status_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"🔧 *AI:* Groq + Gemini + OpenRouter\n✅ Operational, {get_title(update.effective_user.id)}.", parse_mode='Markdown')

async def forget_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    conversation_memory[user_id] = []
    save_memory_to_disk()
    await update.message.reply_text(f"🧠 Memory cleared, {get_title(user_id)}.", parse_mode='Markdown')

async def remind(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not context.args or len(context.args) < 2:
        await update.message.reply_text("⏰ `/remind 30min Call John`", parse_mode='Markdown'); return
    time_str = context.args[0].lower()
    task = " ".join(context.args[1:])
    minutes = 0
    if "min" in time_str: minutes = int(time_str.replace("min","").replace("s","").strip() or 1)
    elif "h" in time_str or "hour" in time_str:
        t = time_str.replace("h","").replace("hour","").replace("hours","").replace("s","").strip()
        minutes = int(t or 1) * 60
    else: await update.message.reply_text("❌ Use: 30min, 1h, 2hour"); return
    rt = datetime.now() + timedelta(minutes=minutes)
    get_user_data(user_id)["reminders"].append({"task": task, "time": rt.strftime("%I:%M %p"), "date": rt.strftime("%b %d")})
    await update.message.reply_text(f"✅ *Reminder:* {task}\n⏰ {rt.strftime('%I:%M %p, %B %d')}", parse_mode='Markdown')

async def view_reminders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    udata = get_user_data(update.effective_user.id)
    if not udata["reminders"]: await update.message.reply_text("📭 No reminders."); return
    rlist = "\n".join([f"{i}. ⏰ {r['time']} — {r['task']}" for i, r in enumerate(udata["reminders"], 1)])
    await update.message.reply_text(f"*📋 Reminders:*\n\n{rlist}", parse_mode='Markdown')

async def take_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args: await update.message.reply_text("📝 `/note Your text`", parse_mode='Markdown'); return
    get_user_data(update.effective_user.id)["notes"].append({"text": " ".join(context.args), "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M")})
    await update.message.reply_text("📌 Noted!", parse_mode='Markdown')

async def view_notes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    udata = get_user_data(update.effective_user.id)
    if not udata["notes"]: await update.message.reply_text("📭 No notes."); return
    nlist = "\n".join([f"{i}. [{n['timestamp']}] {n['text']}" for i, n in enumerate(udata["notes"], 1)])
    await update.message.reply_text(f"*📒 Notes:*\n\n{nlist}", parse_mode='Markdown')

async def calculate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args: await update.message.reply_text("🧮 `/calc 156*23.5`", parse_mode='Markdown'); return
    try:
        result = eval(" ".join(context.args), {"__builtins__": None}, {"abs": abs, "round": round})
        await update.message.reply_text(f"🧮 = *{result:,}*", parse_mode='Markdown')
    except: await update.message.reply_text("❌ Invalid.")

async def search_web(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args: await update.message.reply_text("🔍 `/search query`", parse_mode='Markdown'); return
    await update.message.reply_text(f"🔍 [Search](https://www.google.com/search?q={'+'.join(context.args)})", parse_mode='Markdown', disable_web_page_preview=False)

async def current_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    now = datetime.now()
    await update.message.reply_text(f"🕐 *{now.strftime('%I:%M:%S %p')}*\n📅 *{now.strftime('%A, %B %d, %Y')}*", parse_mode='Markdown')

async def user_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    udata = get_user_data(user_id)
    total = len(conversation_memory.get(user_id, []))
    await update.message.reply_text(
        f"📄 Docs: {udata['documents_analyzed']}\n📝 Notes: {len(udata['notes'])}\n⏰ Reminders: {len(udata['reminders'])}\n🧠 Messages: {total}\n💎 Premium: {'Yes' if is_premium(user_id) else 'No'}\n👤 {get_title(user_id)} {get_name(user_id)}",
        parse_mode='Markdown'
    )

async def fullreset_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    conversation_memory[user_id] = []
    user_preferences[user_id] = {}
    user_data[user_id] = {"reminders": [], "notes": [], "documents_analyzed": 0, "joined": datetime.now().strftime("%Y-%m-%d %H:%M")}
    save_memory_to_disk()
    await update.message.reply_text("🗑 Complete reset. I now know nothing about you.", parse_mode='Markdown')    

async def clear_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_data[user_id] = {"reminders": [], "notes": [], "documents_analyzed": 0, "joined": datetime.now().strftime("%Y-%m-%d %H:%M")}
    await update.message.reply_text(f"🗑 Cleared, {get_title(user_id)}.", parse_mode='Markdown')

# ============================================
# PREMIUM COMMANDS
# ============================================

async def premium_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "💎 *Jarvis Premium*\n\n📦 Basic — $49\n🚀 Business — $149\n🏢 Enterprise — $299\n\nReply: `buy basic`, `buy business`, `buy enterprise`",
        parse_mode='Markdown'
    )

async def buy_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if is_premium(user_id): await update.message.reply_text("✅ Already premium!"); return
    if not context.args: await update.message.reply_text("Usage: `buy basic`"); return
    plan = context.args[0].lower()
    if plan not in PAYMENT_LINKS: await update.message.reply_text("❌ Choose: basic, business, enterprise"); return
    invoice = f"JARVIS-{user_id}-{datetime.now().strftime('%Y%m%d%H%M')}"
    await update.message.reply_text(
        f"💳 *{plan.upper()} — ${PREMIUM_PRICES[plan]}*\n📋 Invoice: `{invoice}`\n\n[Pay Here]({PAYMENT_LINKS[plan]})\n\nAfter payment: `/activate {invoice}`",
        parse_mode='Markdown', disable_web_page_preview=False
    )

async def activate_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args: await update.message.reply_text("Usage: `/activate INVOICE`"); return
    premium_users.add(update.effective_user.id)
    save_memory_to_disk()
    await update.message.reply_text(f"✅ Premium activated! Welcome, {get_title(update.effective_user.id)}! 🎉\n/features to see what's new.")

async def features_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if is_premium(update.effective_user.id):
        await update.message.reply_text("*✅ PREMIUM*\n\n📄 Documents\n📸 Images\n🎤 Voice\n📊 CSV\n📧 Email\n🗓 Calendar\n⚡ Priority AI", parse_mode='Markdown')
    else:
        await update.message.reply_text("*🆓 FREE*\n\n💬 Chat\n📝 Notes\n⏰ Reminders\n🧮 Calculator\n🔍 Search\n\nUpgrade: /premium", parse_mode='Markdown')

# ============================================
# CALENDAR
# ============================================

async def calendar_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("🗓 `/calendar add Meeting tomorrow 3pm`\n`/calendar list`", parse_mode='Markdown'); return
    action = context.args[0].lower()
    udata = get_user_data(update.effective_user.id)
    if action == "add" and len(context.args) >= 2:
        event = " ".join(context.args[1:])
        udata["reminders"].append({"task": f"📅 {event}", "time": "See description", "date": "Calendar"})
        await update.message.reply_text(f"📅 Added: _{event}_", parse_mode='Markdown')
    elif action in ["list", "today"]:
        events = [r for r in udata["reminders"] if r["task"].startswith("📅")]
        if events: await update.message.reply_text("*🗓 Events:*\n\n" + "\n".join([f"• {e['task'].replace('📅 ','')}" for e in events]), parse_mode='Markdown')
        else: await update.message.reply_text("📭 No events.")

# ============================================
# EMAIL
# ============================================

async def email_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args or len(context.args) < 3:
        await update.message.reply_text("📧 `/email to@email.com Subject Your message`", parse_mode='Markdown'); return
    to_email = context.args[0]
    subject = context.args[1]
    body = " ".join(context.args[2:])
    draft = jarvis_think(update.effective_user.id, f"Write a professional email. To: {to_email}. Subject: {subject}. Message: {body}")
    await update.message.reply_text(f"📧 *Draft:*\n\n*To:* {to_email}\n*Subject:* {subject}\n\n{draft}", parse_mode='Markdown')

# ============================================
# ADMIN
# ============================================

async def admin_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_USER_ID: await update.message.reply_text("❌ Admin only."); return
    msg = " ".join(context.args)
    if not msg: await update.message.reply_text("Usage: /broadcast message"); return
    sent = 0
    for uid in conversation_memory:
        try:
            await context.bot.send_message(chat_id=int(uid), text=f"📢 {msg}")
            sent += 1
        except: pass
    await update.message.reply_text(f"✅ Sent to {sent} users.")

async def admin_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_USER_ID: await update.message.reply_text("❌ Admin only."); return
    await update.message.reply_text(f"👥 {len(conversation_memory)} users.")

# ============================================
# CSV
# ============================================

async def handle_csv_request(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_message = update.message.text
    title = get_title(user_id)
    msg_lower = user_message.lower()
    await update.message.reply_text(f"📊 Generating CSV, {title}...")
    # await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    csv_content = jarvis_think(user_id, f"Generate CSV for: {user_message}. Return ONLY raw CSV, no markdown.")
    csv_content = csv_content.strip()
    if csv_content.startswith("```"):
        csv_content = "\n".join([l for l in csv_content.split("\n") if not l.startswith("```")])
    filename = "data.csv"
    if "budget" in msg_lower: filename = "budget_template.csv"
    elif "expense" in msg_lower: filename = "expenses.csv"
    elif "schedule" in msg_lower: filename = "schedule.csv"
    filepath = f"data/{filename}"
    os.makedirs("data", exist_ok=True)
    with open(filepath, 'w', encoding='utf-8') as f: f.write(csv_content)
    with open(filepath, 'rb') as f:
        await update.message.reply_document(document=f, filename=filename, caption=f"📊 Here's your CSV, {title}.")
    os.remove(filepath)

# ============================================
# DOCUMENTS
# ============================================

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    udata = get_user_data(user_id)
    document = update.message.document
    file_name = document.file_name
    file_size = document.file_size / 1024
    title = get_title(user_id)
    status_msg = await update.message.reply_text(f"📄 Analyzing {file_name}...", parse_mode='Markdown')
    # await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    try:
        file = await document.get_file()
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file_name)[1]) as tmp:
            await file.download_to_drive(tmp.name)
            file_path = tmp.name
        content = ""
        file_ext = os.path.splitext(file_name)[1].lower()
        if file_ext == '.txt':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f: content = f.read()
        elif file_ext == '.pdf':
            from PyPDF2 import PdfReader
            for i, page in enumerate(PdfReader(file_path).pages):
                t = page.extract_text()
                if t: content += f"\n--- Page {i+1} ---\n{t}"
        elif file_ext in ['.docx', '.doc']:
            from docx import Document
            content = "\n".join([p.text for p in Document(file_path).paragraphs if p.text.strip()])
        elif file_ext == '.csv':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                rows = list(csv.reader(f))
                content = f"CSV: {len(rows)} rows\n" + "\n".join([" | ".join(row) for row in rows[:50]])
        elif file_ext in ['.xlsx', '.xls']:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            content = f"Excel: {len(wb.sheetnames)} sheets\n"
            for sn in wb.sheetnames[:3]:
                ws = wb[sn]
                content += f"\n--- {sn} ---\n"
                for row in ws.iter_rows(values_only=True, max_row=20):
                    content += " | ".join([str(c) if c else "" for c in row]) + "\n"
        else:
            content = f"[Unsupported: {file_ext}]"
        udata["documents_analyzed"] += 1
        add_to_memory(user_id, "user", f"[Uploaded: {file_name}]")
        if content and len(content) > 50:
            analysis = jarvis_think(user_id, f"Analyze this document:\n\n{content[:5000]}")
            await status_msg.edit_text(f"*📄 {file_name}*\n📊 {file_size:.1f} KB\n\n{analysis}\n\n💡 Ask me about this document, {title}!", parse_mode='Markdown')
        else:
            await status_msg.edit_text(f"⚠️ Could not extract text from {file_name}")
    except Exception as e:
        await status_msg.edit_text(f"❌ Error: {str(e)[:200]}")
    finally:
        if 'file_path' in locals(): os.unlink(file_path)

# ============================================
# PDF GENERATION
# ============================================

async def pdf_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Generate a PDF document"""
    user_id = update.effective_user.id
    title = get_title(user_id) or "there"
    
    if not context.args:
        await update.message.reply_text(
            "📄 *PDF Generator*\n\n"
            "Usage: `/pdf invoice John Doe $149`\n"
            "Usage: `/pdf report Monthly Sales Summary`\n"
            "Usage: `/pdf letter Thank you for your purchase`",
            parse_mode='Markdown'
        )
        return
    
    doc_type = context.args[0].lower()
    content = " ".join(context.args[1:])
    
    await update.message.reply_text(f"📄 Generating PDF, {title}...")
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    
    # Ask AI to generate the content
    prompt = f"Generate a professional {doc_type} document. Content: {content}. Keep it concise and well-formatted."
    pdf_content = jarvis_think(user_id, prompt)
    
    # Create PDF using reportlab (lightweight)
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        import io as io_module
        
        buffer = io_module.BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        
        # Write content
        y = 750
        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, y, f"J.A.R.V.I.S. - {doc_type.upper()}")
        y -= 30
        
        c.setFont("Helvetica", 10)
        c.drawString(50, y, f"Date: {datetime.now().strftime('%B %d, %Y')}")
        y -= 20
        
        c.setFont("Helvetica", 11)
        for line in pdf_content.split('\n'):
            if y < 50:
                c.showPage()
                c.setFont("Helvetica", 11)
                y = 750
            # Wrap long lines
            while len(line) > 90:
                c.drawString(50, y, line[:90])
                y -= 15
                line = line[90:]
            c.drawString(50, y, line)
            y -= 15
        
        c.save()
        buffer.seek(0)
        
        filename = f"{doc_type}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        await update.message.reply_document(
            document=buffer,
            filename=filename,
            caption=f"📄 Here's your {doc_type}, {title}!"
        )
        
    except ImportError:
        await update.message.reply_text("📄 PDF support: `pip install reportlab`")
    except Exception as e:
        await update.message.reply_text(f"❌ PDF error: {str(e)[:200]}")

# ============================================
# TRANSLATION
# ============================================

async def translate_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Translate text to any language"""
    user_id = update.effective_user.id
    title = get_title(user_id) or "there"
    
    if not context.args or len(context.args) < 2:
        await update.message.reply_text(
            "🌍 *Translator*\n\n"
            "Usage: `/translate french Hello, how are you?`\n"
            "Usage: `/translate spanish Good morning`\n"
            "Usage: `/translate japanese Thank you`",
            parse_mode='Markdown'
        )
        return
    
    target_lang = context.args[0].capitalize()
    text = " ".join(context.args[1:])
    
    await update.message.reply_text(f"🌍 Translating to {target_lang}...")
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    
    prompt = f"Translate the following text to {target_lang}. Return ONLY the translation, nothing else:\n\n{text}"
    translation = jarvis_think(user_id, prompt)
    
    await update.message.reply_text(
        f"🌍 *Translation ({target_lang})*\n\n"
        f"📝 Original: _{text}_\n"
        f"🔄 Translated: *{translation}*",
        parse_mode='Markdown'
    )
    
# ============================================
# YOUTUBE VIDEO SUMMARIZER
# ============================================

async def handle_youtube_link(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Detect YouTube links and summarize the video"""
    user_id = update.effective_user.id
    user_message = update.message.text
    title = get_title(user_id) or "there"
    
        # Check if message contains a YouTube link
    youtube_patterns = [
        r'(?:youtube\.com/watch\?v=|youtu\.be/|youtube\.com/shorts/)([a-zA-Z0-9_\-]+)'
    ]
    
    video_id = None
    for pattern in youtube_patterns:
        match = re.search(pattern, user_message)
        if match:
            video_id = match.group(1)
            # Strip any extra parameters after the ID
            if '?' in video_id:
                video_id = video_id.split('?')[0]
            if '&' in video_id:
                video_id = video_id.split('&')[0]
            break
    
    if not video_id:
        return False  # Not a YouTube link
    
    await update.message.reply_text(f"📺 Analyzing YouTube video, {title}...")
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    
    try:
        # Get video info using oEmbed (no API key needed!)
        oembed_url = f"https://www.youtube.com/oembed?url=https://www.youtube.com/watch?v={video_id}&format=json"
        response = requests.get(oembed_url, timeout=10)
        
        if response.status_code == 200:
            video_data = response.json()
            video_title = video_data.get('title', 'Unknown Title')
            author = video_data.get('author_name', 'Unknown Creator')
        else:
            video_title = "YouTube Video"
            author = "Unknown"
        
        # Ask AI to summarize based on the title and context
        prompt = f"""A user wants a summary of this YouTube video:
        Title: {video_title}
        Creator: {author}
        URL: https://www.youtube.com/watch?v={video_id}
        
        Based on the title, provide:
        1. What this video is likely about (2-3 sentences)
        2. Key points the video probably covers (3-5 bullet points)
        3. Who would benefit from watching this
        4. Estimated time saved by reading this summary
        
        Be honest that this is a title-based summary. Suggest they can share more context for a better summary."""
        
        summary = jarvis_think(user_id, prompt)
        
        await update.message.reply_text(
            f"📺 *YouTube Summary*\n\n"
            f"🎬 *{video_title}*\n"
            f"👤 {author}\n\n"
            f"{summary}\n\n"
            f"💡 *Tip:* Tell me more about the video for a better summary!\n"
            f"Example: 'This video is about Python tutorials for beginners'",
            parse_mode='Markdown'
        )
        
        add_to_memory(user_id, "user", user_message)
        add_to_memory(user_id, "assistant", f"Summarized YouTube video: {video_title}")
        return True
        
    except Exception as e:
        await update.message.reply_text(f"❌ Could not fetch video info. Try pasting the video title instead.")
        return True  # Still handled

async def summarize_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Summarize any YouTube video by title/topic"""
    user_id = update.effective_user.id
    title = get_title(user_id) or "there"
    
    if not context.args:
        await update.message.reply_text(
            "📺 *YouTube Summarizer*\n\n"
            "Send me a YouTube link OR use:\n"
            "`/summarize Video title or topic`\n\n"
            "Example: `/summarize Python tutorial for beginners`",
            parse_mode='Markdown'
        )
        return
    
    topic = " ".join(context.args)
    
    await update.message.reply_text(f"📺 Summarizing '{topic}'...")
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    
    prompt = f"Provide a helpful summary about this YouTube video topic: '{topic}'. Include what viewers can expect to learn, key takeaways, and who this content is for."
    
    summary = jarvis_think(user_id, prompt)
    
    await update.message.reply_text(
        f"📺 *Video Summary: {topic}*\n\n{summary}",
        parse_mode='Markdown'
    )

# ============================================
# ROAST ME / PROFILE ANALYZER
# ============================================

async def roast_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Roast or analyze based on user's chat history"""
    user_id = update.effective_user.id
    title = get_title(user_id) or "friend"
    
    # Get user's conversation history
    history = get_recent_memory(user_id, 15)
    
    if not history or history == "New conversation":
        await update.message.reply_text(
            "🔍 *Not enough data to roast you yet!*\n\n"
            "Chat with me more, then try again.\n"
            "Or send a self-description: `/roast I'm a 25-year-old software developer who loves pizza and procrastination`",
            parse_mode='Markdown'
        )
        return
    
    await update.message.reply_text(f"🔥 Analyzing your personality, {title}...")
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    
    # If user provided self-description in the command
    extra_info = " ".join(context.args) if context.args else ""
    
    prompt = f"""Based on this conversation history, create a FUNNY and WITTY roast/analysis of the user. 
    Be humorous but NOT mean. Include:
    
    1. 🔥 A funny roast (1-2 sentences, playful)
    2. 🧠 Personality traits you detect (3-4)
    3. 💼 Career they might have
    4. 🎯 Hidden talent they probably have
    5. ⭐ Overall rating (out of 10)
    
    Extra info from user: {extra_info}
    
    Conversation history:
    {history}
    
    Make it funny and shareable. Use emojis. End with something encouraging."""
    
    roast = jarvis_think(user_id, prompt)
    
    await update.message.reply_text(
        f"🔥 *Roast Mode Activated*\n\n{roast}\n\n"
        f"_Send this to a friend! They can try too: just chat with me then type /roast_",
        parse_mode='Markdown'
    )

async def analyze_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """More serious personality analysis"""
    user_id = update.effective_user.id
    title = get_title(user_id) or "friend"
    
    history = get_recent_memory(user_id, 15)
    
    if not history or history == "New conversation":
        await update.message.reply_text("🔍 Chat more first, then I can analyze you!")
        return
    
    await update.message.reply_text(f"🧠 Analyzing, {title}...")
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    
    prompt = f"""Based on this conversation history, provide a PROFESSIONAL personality analysis:
    
    1. MBTI type guess (e.g., INTJ, ENFP)
    2. Communication style
    3. Strengths (3-4)
    4. Growth areas (2-3, kindly worded)
    5. Ideal career matches
    
    Conversation:
    {history}"""
    
    analysis = jarvis_think(user_id, prompt)
    
    await update.message.reply_text(f"🧠 *Personality Analysis*\n\n{analysis}", parse_mode='Markdown')

    
# ============================================
# IMAGE ANALYSIS (UPDATED — uses google-genai)
# ============================================

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    title = get_title(user_id)
    
    if not GEMINI_API_KEY or GEMINI_API_KEY == "YOUR_GEMINI_KEY_HERE":
        await update.message.reply_text(f"📸 Image analysis requires Gemini API key, {title}. Get one at https://aistudio.google.com/apikey")
        return
    
    await update.message.reply_text(f"📸 Analyzing image, {title}...")
    # await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    
    try:
        from google import genai
        from google.genai import types
        
        client = genai.Client(api_key=GEMINI_API_KEY)
        
        photo = update.message.photo[-1]
        file = await photo.get_file()
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
            await file.download_to_drive(tmp.name)
            image_path = tmp.name
        
        # Upload image to Gemini
        with open(image_path, 'rb') as img:
            image_bytes = img.read()
        
        caption = update.message.caption or "Describe this image in detail."
        
        response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=[
                types.Part.from_bytes(data=image_bytes, mime_type="image/jpeg"),
                caption
            ]
        )
        
        await update.message.reply_text(
            f"📸 *Analysis:*\n\n{response.text}\n\n💡 *Tip:* Add a caption like 'Read the text in this image' for OCR!",
            parse_mode='Markdown'
        )
        
        os.unlink(image_path)
        
    except ImportError:
        await update.message.reply_text("📸 Run: `pip install google-genai`")
    except Exception as e:
        await update.message.reply_text(f"❌ Image analysis failed: {str(e)[:200]}")

# ============================================
# VOICE TRANSCRIPTION (NEW!)
# ============================================

async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    title = get_title(user_id)
    
    await update.message.reply_text(f"🎤 Transcribing, {title}...")
    # await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    
    try:
        voice = update.message.voice
        file = await voice.get_file()
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".ogg") as tmp:
            await file.download_to_drive(tmp.name)
            audio_path = tmp.name
        
        url = "https://api.groq.com/openai/v1/audio/transcriptions"
        with open(audio_path, 'rb') as af:
            response = requests.post(
                url,
                headers={"Authorization": f"Bearer {GROQ_API_KEY}"},
                files={'file': ('audio.ogg', af, 'audio/ogg'), 'model': (None, 'whisper-large-v3')},
                timeout=30
            )
        
        if response.status_code == 200:
            transcript = response.json()["text"]
            await update.message.reply_text(f"🎤 *You said:* _{transcript}_\n\nProcessing...", parse_mode='Markdown')
            
            add_to_memory(user_id, "user", transcript)
            ai_resp = jarvis_think(user_id, transcript)
            add_to_memory(user_id, "assistant", ai_resp)
            
            if len(ai_resp) > 4000:
                for i in range(0, len(ai_resp), 4000):
                    await update.message.reply_text(ai_resp[i:i+4000])
            else:
                await update.message.reply_text(ai_resp)
        else:
            await update.message.reply_text(f"❌ Transcription failed. Error: {response.status_code}")
        
        os.unlink(audio_path)
    except Exception as e:
        await update.message.reply_text(f"❌ Voice error: {str(e)[:200]}")

# ============================================
# CHAT
# ============================================

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_message = update.message.text
    title = get_title(user_id)
    
    # await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    msg_lower = user_message.lower().strip()
    
    time_only = ["time", "date", "what time is it", "what's the time", "current time", "today's date", "what day is it", "what day is today", "what is the time", "tell me the time"]
    if msg_lower in time_only:
        await current_time(update, context); return
    
    # YouTube link detection
    if await handle_youtube_link(update, context):
        return
    
    csv_triggers = ["give me a csv", "create a csv", "generate a csv", "make a csv", "csv format", "csv file", "excel file", "spreadsheet file", "generate csv", "create csv", "make csv", "send me a csv", "download csv", ".csv"]
    if any(t in msg_lower for t in csv_triggers):
        await handle_csv_request(update, context); return
    
    if detect_user_info(user_message, user_id):
        title = get_title(user_id)
        name = get_name(user_id)
        await update.message.reply_text(f"✅ Noted, {title} {name}!" if name else f"✅ Understood, {title}!", parse_mode='Markdown')
        add_to_memory(user_id, "user", user_message); return
    
    if is_affirmation(user_message) and conversation_memory.get(user_id):
        for msg in reversed(conversation_memory[user_id]):
            if msg["role"] == "user":
                user_message = f"User said '{user_message}'. Elaborate on our last topic."; break
    
    add_to_memory(user_id, "user", user_message)
    
    # Get AI response with longer timeout
    try:
        response = jarvis_think(user_id, user_message)
        add_to_memory(user_id, "assistant", response)
        
        if len(response) > 4000:
            for i in range(0, len(response), 4000):
                await update.message.reply_text(response[i:i+4000], read_timeout=30, write_timeout=30)
        else:
            await update.message.reply_text(response, read_timeout=30, write_timeout=30)
    except Exception as e:
        await update.message.reply_text("⚠️ Response took too long. Please try again.")

# ============================================
# MAIN
# ============================================

def main():
    print("""
╔══════════════════════════════════╗
║     J.A.R.V.I.S. v3.0 ONLINE     ║
║   AI Chat | Images | Voice       ║
║   Docs | CSV | Email | Calendar  ║
║   Premium Plans Ready            ║
╚══════════════════════════════════╝
    """)
    
    load_memory_from_disk()
    
    request = HTTPXRequest(connection_pool_size=8, read_timeout=60, write_timeout=60, connect_timeout=60)
    app = Application.builder().token(TELEGRAM_TOKEN).request(request).build()
    
    # Basic
       # Basic
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("status", status_command))
    app.add_handler(CommandHandler("whoami", whoami_command))
    app.add_handler(CommandHandler("history", history_command))
    app.add_handler(CommandHandler("forget", forget_command))
    app.add_handler(CommandHandler("fullreset", fullreset_command))
    app.add_handler(CommandHandler("clear", clear_data))  

    
    
    # Productivity
    app.add_handler(CommandHandler("remind", remind))
    app.add_handler(CommandHandler("reminders", view_reminders))
    app.add_handler(CommandHandler("note", take_note))
    app.add_handler(CommandHandler("notes", view_notes))
    app.add_handler(CommandHandler("calc", calculate))
    app.add_handler(CommandHandler("search", search_web))
    app.add_handler(CommandHandler("time", current_time))
    app.add_handler(CommandHandler("stats", user_stats))
    app.add_handler(CommandHandler("clear", clear_data))
    
    # Premium
    app.add_handler(CommandHandler("premium", premium_command))
    app.add_handler(CommandHandler("buy", buy_command))
    app.add_handler(CommandHandler("activate", activate_command))
    app.add_handler(CommandHandler("features", features_command))
    
    # Calendar & Email
    app.add_handler(CommandHandler("calendar", calendar_command))
    app.add_handler(CommandHandler("email", email_command))
    
    # Admin
    app.add_handler(CommandHandler("broadcast", admin_broadcast))
    app.add_handler(CommandHandler("users", admin_users))

    # Fun & Viral
    app.add_handler(CommandHandler("roast", roast_command))
    app.add_handler(CommandHandler("analyze", analyze_command))
    app.add_handler(CommandHandler("summarize", summarize_command))
    
    # Files & Media
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.VOICE, handle_voice))
    
    # Chat
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.add_handler(CommandHandler("receipt", receipt_command))
    print("✅ Online with ALL features!")
    print("📸 Send a photo | 🎤 Send a voice note | 📄 Upload a document")
    print("💰 /premium to see plans\n")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

    # ============================================
# RECEIPT GENERATOR (Generates downloadable receipt)
# ============================================

async def receipt_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Generate a payment receipt"""
    user_id = update.effective_user.id
    title = get_title(user_id)
    name = get_name(user_id) or "Valued Customer"
    
    if not context.args:
        await update.message.reply_text(
            "🧾 *Receipt Generator*\n\n"
            "Usage: `/receipt [amount] [item]`\n"
            "Example: `/receipt 149 Business Plan`\n"
            "Example: `/receipt 49 Basic Subscription`",
            parse_mode='Markdown'
        )
        return
    
    amount = context.args[0]
    item = " ".join(context.args[1:]) if len(context.args) > 1 else "Service"
    
    # Generate receipt number
    receipt_num = f"JAR-{datetime.now().strftime('%Y%m%d')}-{user_id}"
    invoice_num = f"INV-{datetime.now().strftime('%Y%m%d%H%M')}"
    date_str = datetime.now().strftime("%B %d, %Y")
    time_str = datetime.now().strftime("%I:%M %p")
    
    # Create receipt as text file
    receipt_text = f"""
╔══════════════════════════════════════════╗
║         J.A.R.V.I.S. TECHNOLOGIES        ║
║           PAYMENT RECEIPT                 ║
╚══════════════════════════════════════════╝

Date: {date_str}
Time: {time_str}
Receipt #: {receipt_num}
Invoice #: {invoice_num}

──────────────────────────────────────────
CUSTOMER DETAILS
──────────────────────────────────────────
Name: {name}
Customer ID: {user_id}
Status: {'💎 Premium' if is_premium(user_id) else '🆓 Free'}

──────────────────────────────────────────
PAYMENT DETAILS
──────────────────────────────────────────
Item: {item}
Amount: ${amount}
Payment Method: Online Payment
Status: ✅ PAID

──────────────────────────────────────────
Thank you for your purchase, {title} {name}!
For support: @YourSupportUsername
──────────────────────────────────────────

This is a computer-generated receipt.
"""
    
    # Save to file
    filename = f"receipt_{receipt_num}.txt"
    filepath = f"data/{filename}"
    os.makedirs("data", exist_ok=True)
    
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(receipt_text)
    
    # Send the receipt
    with open(filepath, 'rb') as f:
        await update.message.reply_document(
            document=f,
            filename=filename,
            caption=f"🧾 Here's your receipt for ${amount}, {title} {name}!"
        )
    
    os.remove(filepath)

if __name__ == "__main__":
    main()